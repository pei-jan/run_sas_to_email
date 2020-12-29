#!/usr/bin/env python
# coding: utf-8

# In[3]:


def email表格():
    import openpyxl , xlrd
    wb = openpyxl.load_workbook('轉存mail.xlsm')
    ws = wb.active
    #開頭文字
    contents_1 = '''
    <p><strong><span style="color:#003060"><span style="font-size:24px"><font face='Microsoft JhengHei'>Dear all：</font></span></span></p>
    <p></p>
    <p><u><strong><span style="color:#EA0000"><span style="font-size:24px"><font face='Microsoft JhengHei'>請輔導經理務必協助專展同仁審視要保文件！以減少不全率。</font></span></span></u></p>

    <p><strong><span style="color:#EA0000"><span style="font-size:24px"><font face='Microsoft JhengHei'>
    '''+ws['A16'].value+'''
    '''+ws['A17'].value+'''
    '''+'掌聲鼓勵以上同仁~</font></span></span></p>'+'''
    <p><strong><span style="color:#003060"><span style="font-size:24px"><font face='Microsoft JhengHei'>'''+ws['A18'].value+'''
    '''+'</font></span></span></p>'+'''
    <p><strong><span style="color:#BB5E00"><span style="font-size:24px"><font face='Microsoft JhengHei'>'''+ws['A19'].value+'''
    </font></span></span></p>'''+'''
    <p><strong><span style="color:#003060"><span style="font-size:24px"><font face='Microsoft JhengHei'>'''+ws['A20'].value+'''
    </font></span></span></p>'''+'''
    <p><strong><span style="color:#EA0000"><span style="font-size:24px"><font face='Microsoft JhengHei'>'''+ws['A24'].value+'''
    </font></span></span></p>
    '''
    #團營表格
    a = 0
    book = xlrd.open_workbook('轉存mail.xlsm')
    sheet = book.sheet_by_index(0)
    for col in sheet.col_values(6):
        if col != '':
            a+=1
    contents_2 = '''
    <table align="center" border="5" cellpadding="5" cellspacing="1" style="border: 3px bgcolor= #FFFFFF solid #FFAE7F; 
    border-collapse: collapse; width:1200px; "><font face='Microsoft JhengHei'>
    <tbody>
     <tr>
      <td width="250" valign="middle" align='center' bgcolor= #fff082 >排名</td>
      <td width="450" valign="middle" align='center' bgcolor= #fff082>團營單位</td>
      <td width="450" valign="middle" align='center' bgcolor= #fff082>團營經理</td>
      <td width="400" valign="middle" align='center' bgcolor= #fff082>達成率-套裝初實收(初年度)</td>
      <td width="400" valign="middle" align='center' bgcolor= #fff082>達成率-初實收(初年度)</td>
      <td width="850" valign="middle" align='center' bgcolor= #fff082>累計進度</td>
      <td width="250" valign="middle" align='center' bgcolor= #fff082>排名</td>
      <td width="450" valign="middle" align='center' bgcolor= #fff082>團營單位</td>
      <td width="450" valign="middle" align='center' bgcolor= #fff082>團營經理</td>
      <td width="320" valign="middle" align='center' bgcolor= #fff082>達成率-總實收(初+續)</td>
      <td width="850" valign="middle" align='center' bgcolor= #fff082>累計進度</td>
     '''
    for i in range(a-2):
        contents_2 += '''
         <tr>
        <td valign="middle" align='center'>
        '''+ str(ws['A'+str(i+3)].value) +'''
        </td>
        <td valign="middle" align='center'>
        '''+ str(ws['B'+str(i+3)].value) +'''
        </td>
        <td valign="middle" align='center'>
        '''+ str(ws['C'+str(i+3)].value) +'''
        </td>
        <td valign="middle" align='center'>
        '''+ '%.2f'%(ws["D"+str(i+3)].value*100)+'%' +'''
        </td>
        <td valign="middle" align='center'>
        '''+ '%.2f'%(ws["E"+str(i+3)].value*100)+'%' +'''
        </td>
        <td valign="middle" align='center'>
        '''+ str(ws['F'+str(i+3)].value) +'''
        </td>
        <td valign="middle" align='center'>
        '''+ str(ws['G'+str(i+3)].value) +'''
        </td>
        <td valign="middle" align='center'>
        '''+ str(ws['H'+str(i+3)].value) +'''
        </td>
        <td valign="middle" align='center'>
        '''+ str(ws['I'+str(i+3)].value) +'''
        </td>
        <td valign="middle" align='center'>
        '''+ '%.2f'%(ws["J"+str(i+3)].value*100)+'%' +'''
        </td>
        <td valign="middle" align='center'>
        '''+ str(ws['K'+str(i+3)].value) +'''
        </td>
         </tr>
        '''
    contents_2 += '''
    <td valign="middle" align='center'>'''+ str(ws['A'+str(a+1)].value) +'''</td>
    <td valign="middle" align='center'>-</td>
    <td valign="middle" align='center'>-</td>
    <td valign="middle" align='center'>'''+ '%.2f'%(ws['D'+str(a+1)].value*100)+'%' +'''</td>
    <td valign="middle" align='center'>'''+ '%.2f'%(ws['E'+str(a+1)].value*100)+'%' +'''</td>
    <td valign="middle" align='center'>'''+ str(ws['F'+str(a+1)].value) +'''</td>
    <td valign="middle" align='center'>'''+ str(ws['G'+str(a+1)].value) +'''</td>
    <td valign="middle" align='center'>-</td>
    <td valign="middle" align='center'>-</td>
    <td valign="middle" align='center'>'''+ '%.2f'%(ws['J'+str(a+1)].value*100)+'%' +'''</td>
    <td valign="middle" align='center'>'''+ str(ws['K'+str(a+1)].value) +'''</td>
    </tr>
    </font>
    </tbody>
    </table>
    <p>&nbsp;</p>
    '''
    #輔經表格
    b = 0
    for col in sheet.col_values(12):
        if col != '':
            b+=1
    contents_3 = '''
    <table align="center" border="5" cellpadding="5" cellspacing="1" style="border: 3px bgcolor= #FFFFFF solid #FFAE7F; 
    border-collapse: collapse; width:1500px; "><font face='Microsoft JhengHei'>
    <tbody>
     <tr>
      <td width="150" valign="middle" align='center' bgcolor= #fff082 >排名</td>
      <td width="450" valign="middle" align='center' bgcolor= #fff082>團營單位</td>
      <td width="700" valign="middle" align='center' bgcolor= #fff082>輔導經理</td>
      <td width="700" valign="middle" align='center' bgcolor= #fff082>團營經理</td>
      <td width="400" valign="middle" align='center' bgcolor= #fff082>達成率-套裝初實收(初年度)</td>
      <td width="400" valign="middle" align='center' bgcolor= #fff082>達成率-初實收(初年度)</td>
      <td width="750" valign="middle" align='center' bgcolor= #fff082>累計進度</td>
      <td width="150" valign="middle" align='center' bgcolor= #fff082>排名</td>
      <td width="450" valign="middle" align='center' bgcolor= #fff082>團營單位</td>
      <td width="700" valign="middle" align='center' bgcolor= #fff082>輔導經理</td>
      <td width="700" valign="middle" align='center' bgcolor= #fff082>團營經理</td>
      <td width="320" valign="middle" align='center' bgcolor= #fff082>達成率-總實收(初+續)</td>
      <td width="750" valign="middle" align='center' bgcolor= #fff082>累計進度</td>
     '''
    for i in range(b-2):
        contents_3 += '''
         <tr>
        <td valign="middle" align='center'>
        '''+ str(ws['M'+str(i+3)].value) +'''
        </td>
        <td valign="middle" align='center'>
        '''+ str(ws['N'+str(i+3)].value) +'''
        </td>
        <td valign="middle" align='center'>
        '''+ str(ws['O'+str(i+3)].value) +'''
        </td>
        <td valign="middle" align='center'>
        '''+ str(ws['P'+str(i+3)].value) +'''
        </td>
        <td valign="middle" align='center'>
        '''+ '%.2f'%(ws["Q"+str(i+3)].value*100)+'%' +'''
        </td>
        <td valign="middle" align='center'>
        '''+ '%.2f'%(ws["R"+str(i+3)].value*100)+'%' +'''
        </td>
        <td valign="middle" align='center'>
        '''+ str(ws['S'+str(i+3)].value) +'''
        </td>
        <td valign="middle" align='center'>
        '''+ str(ws['T'+str(i+3)].value) +'''
        </td>
        <td valign="middle" align='center'>
        '''+ str(ws['U'+str(i+3)].value) +'''
        </td>
        <td valign="middle" align='center'>
        '''+ str(ws['V'+str(i+3)].value) +'''
        </td>
        <td valign="middle" align='center'>
        '''+ str(ws['W'+str(i+3)].value) +'''
        </td>
        <td valign="middle" align='center'>
        '''+ '%.2f'%(ws["X"+str(i+3)].value*100)+'%' +'''
        </td>
        <td valign="middle" align='center'>
        '''+ str(ws['Y'+str(i+3)].value) +'''
        </td>
         </tr>
        '''
    contents_3 += '''
    <td valign="middle" align='center'>'''+ str(ws['M'+str(b+1)].value) +'''</td>
    <td valign="middle" align='center'>-</td>
    <td valign="middle" align='center'>-</td>
    <td valign="middle" align='center'>-</td>
    <td valign="middle" align='center'>'''+ '%.2f'%(ws['Q'+str(b+1)].value*100)+'%' +'''</td>
    <td valign="middle" align='center'>'''+ '%.2f'%(ws['R'+str(b+1)].value*100)+'%' +'''</td>
    <td valign="middle" align='center'>'''+ str(ws['S'+str(b+1)].value) +'''</td>
    <td valign="middle" align='center'>'''+ str(ws['T'+str(b+1)].value) +'''</td>
    <td valign="middle" align='center'>-</td>
    <td valign="middle" align='center'>-</td>
    <td valign="middle" align='center'>-</td>
    <td valign="middle" align='center'>'''+ '%.2f'%(ws['X'+str(b+1)].value*100)+'%' +'''</td>
    <td valign="middle" align='center'>'''+ str(ws['Y'+str(b+1)].value) +'''</td>
    </tr>
    </font>
    </tbody>
    </table>
    '''
    contents = contents_1 + contents_2 + contents_3
    return contents

def wait_sas():
    import os , time
    try:
        os.remove(r'D:\SASoutput\FINAL_SALES_51.xlsx')
    except:
        pass
    time.sleep(1)
    list_download = []
    for i in os.listdir(r'D:\SASoutput'):
        list_download.append(i)
    if any('FINAL_SALES_51' in s for s in list_download):
        print('SAS執行完成')
        pass
    else:
        time.sleep(10)
        wait_sas()
def run_sas():
    from win32con import WM_INPUTLANGCHANGEREQUEST
    import win32com.client as win32
    import win32gui , win32api , win32con ,time, pyautogui , glob
    sas_file = glob.glob('*.egp')
    #執行sas檔
    win32api.ShellExecute(0, 'open', sas_file[0], '', '', 1)
    time.sleep(5)
    hwnd = win32gui.GetForegroundWindow()
    win32gui.ShowWindow(hwnd, win32con.SW_MAXIMIZE)

    time.sleep(5)
    pyautogui.click(1000,600)
    time.sleep(1)
    pyautogui.keyDown('alt')
    time.sleep(1)
    pyautogui.press('r')
    pyautogui.keyUp('alt')
    time.sleep(1)
    for _ in range(4):
        pyautogui.press('down')
        time.sleep(0.01)
    pyautogui.press('enter')
    print('開始執行SAS')
def 中英切換():
    from win32con import WM_INPUTLANGCHANGEREQUEST
    import win32gui
    import win32api

    # 語言程式碼
    # https://msdn.microsoft.com/en-us/library/cc233982.aspx
    LID = {0x0804: "Chinese (Simplified) (People's Republic of China)",
           0x0409: 'English (United States)'}

    # 獲取前景視窗控制代碼
    hwnd = win32gui.GetForegroundWindow()

    # 獲取前景視窗標題
    title = win32gui.GetWindowText(hwnd)
    print('當前視窗：' + title)

    # 獲取鍵盤佈局列表
    im_list = win32api.GetKeyboardLayoutList()
    im_list = list(map(hex, im_list))
    print(im_list)

    # 設定鍵盤佈局為英文
    result = win32api.SendMessage(
        hwnd,
        WM_INPUTLANGCHANGEREQUEST,
        0,
        0x0409)
    if result == 0:
        print('設定英文鍵盤成功！')

def run_excel_vba():
    import time
    import xlwings as xw
    import os
    print('開始執行0巨集')
    wb = xw.Book(os.getcwd()+'\\0.每日績效整理.xlsm')
    wb.macro('複製SAS速報報表').run()
    
    time.sleep(3)
    print('開始執行1巨集')
    wb = xw.Book(os.getcwd()+'\\1.輔導組達成率速報.xlsm')
    wb.macro('整理速報資料').run()
    time.sleep(3)
    
    print('開始執行轉存巨集')
    wb = xw.Book(os.getcwd()+'\\轉存mail.xlsm')
    wb.macro('email轉存檔').run()
def move_sasoutput():
    #移動sas產出的檔案
    import os
    import shutil
    try:
        shutil.copy(r'D:\SASoutput\AFFAIR_41_Q1_AG.xlsx',os.getcwd())
        print('move'+ 'AFFAIR_41_Q1_AG.xlsx')
    except:
        pass
    try:
        shutil.copy(r'D:\SASoutput\AFFAIR_41_Q2_AG.xlsx',os.getcwd())
        print('move'+ 'AFFAIR_41_Q2_AG.xlsx')
    except:
        pass
    try:
        shutil.copy(r'D:\SASoutput\AFFAIR_41_Q3_AG.xlsx',os.getcwd())
        print('move'+ 'AFFAIR_41_Q3_AG.xlsx')
    except:
        pass
    try:
        shutil.copy(r'D:\SASoutput\AFFAIR_41_Q4_AG.xlsx',os.getcwd())
        print('move'+ 'AFFAIR_41_Q4_AG.xlsx')
    except:
        pass
    try:
        shutil.copy(r'D:\SASoutput\FILTER_AFFAIR_Q1_AG.xlsx',os.getcwd())
        print('move'+ 'FILTER_AFFAIR_Q1_AG.xlsx')
    except:
        pass
    try:
        shutil.copy(r'D:\SASoutput\FILTER_AFFAIR_Q2_AG.xlsx',os.getcwd())
        print('move'+ 'FILTER_AFFAIR_Q2_AG.xlsx')
    except:
        pass
    try:
        shutil.copy(r'D:\SASoutput\FILTER_AFFAIR_Q3_AG.xlsx',os.getcwd())
        print('move'+ 'FILTER_AFFAIR_Q3_AG.xlsx')
    except:
        pass
    try:
        shutil.copy(r'D:\SASoutput\FILTER_AFFAIR_Q4_AG.xlsx',os.getcwd())
        print('move'+ 'FILTER_AFFAIR_Q4_AG.xlsx')
    except:
        pass
    try:
        shutil.copy(r'D:\SASoutput\ROUTINE_SALES.xlsx',os.getcwd())
        print('move'+ 'ROUTINE_SALES.xlsx')
    except:
        pass
    try:
        shutil.copy(r'D:\SASoutput\UNPAY.xlsx',os.getcwd())
        print('move'+ 'UNPAY.xlsx')
    except:
        pass
    print('完成移動SAS產出檔')
    
    
#附加excel、pdf檔
def 附加檔案():
    global mailContent
    global msg
    import glob
    from email.mime.application import MIMEApplication
    import base64
    excel_file = glob.glob('3.*.xlsx')
    mailContent = MIMEApplication(open(excel_file[0],'rb').read())
    attachname = excel_file[0]
    bs_filename = base64.b64encode(attachname.encode('utf-8'))
    attachname = '=?utf-8?b?' +bs_filename.decode() + '?='
    mailContent.add_header('Content-Disposition', 'attachment', filename=(attachname))
    msg.attach(mailContent)

    #附加pdf檔
    pdf_file1 = glob.glob('*_團營單位*.pdf')[0]
    pdf_file2 = glob.glob('*_輔導經理*.pdf')[0]

    pdf = MIMEApplication(open(pdf_file1, 'rb').read())
    pdf_attachname = pdf_file1
    bs_pdf_filename = base64.b64encode(pdf_attachname.encode('utf-8'))
    pdf_attachname = '=?utf-8?b?' +bs_pdf_filename.decode() + '?='
    pdf.add_header('Content-Disposition', 'attachment', filename=pdf_attachname)
    msg.attach(pdf)

    pdf2 = MIMEApplication(open(pdf_file1, 'rb').read())
    pdf2_attachname = pdf_file2
    bs_pdf2_filename = base64.b64encode(pdf2_attachname.encode('utf-8'))
    pdf2_attachname = '=?utf-8?b?' +bs_pdf_filename.decode() + '?='
    pdf2.add_header('Content-Disposition', 'attachment', filename=pdf_attachname)
    msg.attach(pdf)
    return mailContent , msg

def 收件人mail():
    import openpyxl , xlrd
    wb = openpyxl.load_workbook('收件人mail.xlsx')
    ws = wb.active
    a = 0
    book = xlrd.open_workbook('收件人mail.xlsx')
    sheet = book.sheet_by_index(0)
    for col in sheet.col_values(0):
        if col != '':
            a+=1
    b = 0
    for col in sheet.col_values(2):
        if col != '':
            b+=1
    正本list = []
    for i in range(2,a+1):
        正本list.append(ws['B'+str(i)].value)
    正本收件人 = ','.join(正本list)

    副本list = []
    for i in range(2,b+1):
        副本list.append(ws['D'+str(i)].value)
    副本收件人 = ','.join(副本list)
    return 正本收件人 , 副本收件人

